from flask import Flask, render_template, request
from openpyxl import Workbook, load_workbook
import os
import smtplib
from email.mime.text import MIMEText
from datetime import datetime
app = Flask(__name__)

EXCEL_FILE = "leads.xlsx"


def save_to_excel(name, phone, email):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Phone", "Email", "Date", "Time"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    now = datetime.now()
    date = now.strftime("%d-%m-%Y")
    time = now.strftime("%H:%M:%S")

    ws.append([name, phone, email, date, time])
    wb.save(EXCEL_FILE)
def send_email(name, phone, email):
    sender = "patilsiddh2026@gmail.com"
    password = "qiexjmucdmzayknw"
    receiver = "patil123siddh@gmail.com"

    msg = MIMEText(f"New Lead Received!\n\nName: {name}\nPhone: {phone}\nEmail: {email}")
    msg["Subject"] = "New Real Estate Lead"
    msg["From"] = sender
    msg["To"] = receiver

    server = smtplib.SMTP("smtp.gmail.com", 587)
    server.starttls()
    server.login(sender, password)
    server.sendmail(sender, receiver, msg.as_string())
    server.quit()

@app.route("/")
def home():
    return render_template("sp.html")

@app.route("/submit", methods=["POST"])
def submit():
    name = request.form["name"]
    phone = request.form["phone"]
    email = request.form.get("email", "")

    save_to_excel(name, phone, email)
    send_email(name, phone, email)

    return "Success"

if __name__ == "__main__":
    app.run(debug=True)